import pandas as pd 
import openpyxl
from openai import OpenAI

client = OpenAI(api_key=GPT_API_KEY)

# Create a new Excel workbook
workbookResult = openpyxl.Workbook()

# Select the active sheet (by default, it's the first sheet created)
sheetResult = workbookResult.active

# Initialize the row index for the result sheet
result_row_index = 1

# Load your existing Excel file
workbookInput  = openpyxl.load_workbook('file.xlsx')
sheetInput = workbookInput.active

# Loop through each cell in your Greek data
for row in sheetInput.iter_rows(min_row=2, max_row=sheetInput.max_row, min_col=1, max_col=1):
    if row[0].value != None: 
      # print(row[0].value + "\n")

      response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[
          {
            "role": "system",
            "content": "You will be provided with unstructured data, and your task is to parse it into CSV format. each cell should contain 1 question regarding the process \n\nQuestions must be in format of: What does the process concern about ? Conditions of the process : Steps of the process : \n\nthe word process must contain the name of the process mentioned in the text , for every question there must be an answer , all the question - answer must be in greek , format question - answer :  question answer "
          },
          {
            "role": "user",
            "content": row[0].value
          },
        ],
        temperature=0,
        max_tokens=3755,
        top_p=1,
        frequency_penalty=0,
        presence_penalty=0
      )

      content = response.choices[0].message.content
      # print('content = ', content)

      contents = content.split('\n\n')

      for i, element in enumerate(contents, 1):
        print(f"Paragraph {i}:\n{element}\n")
        sheetResult.cell(row=result_row_index, column=1, value=element)
        result_row_index += 1
      
      # Save the Excel file
      workbookResult.save('outputQuestionsAnswers.xlsx')
       

print(f"done\n")

